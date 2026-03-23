"""
Thesis Project Allocation System — LP Optimisation
====================================================
Uses the Hungarian algorithm (via scipy) to find the globally optimal
assignment of students to projects, minimising total preference rank.

This replicates the GLPK/AMPL approach from the original code:
    minimise sum of cost[i,j] * assign[i,j]
    subject to: each student gets exactly 1 project
                each project goes to at most 1 student

HOW TO USE:
-----------
1. Install dependency (only needed once):
       pip install openpyxl scipy

2. Prepare your Excel file using the provided template:
       thesis_data_TEMPLATE.xlsx
   It must have ONE sheet with these columns:
       Column A: "Student" — student number and name (e.g. "123456 John Smith")
       Column B: "First"   — 1st choice (e.g. "C001 Project Title Here")
       Column C: "Second"  — 2nd choice
       ...up to Column I: "Eighth"

3. Run the script:
       python thesis_allocator_LP.py
   Or specify your file:
       python thesis_allocator_LP.py my_survey_responses.xlsx

4. Output file is saved as:
       thesis_allocation_results.xlsx

COST MATRIX:
------------
  1st choice = cost 1
  2nd choice = cost 2
  ...
  8th choice = cost 8
  Not chosen = cost 1000 (high penalty, last resort only)

The solver minimises the total cost across all students, so it may give
one student their 2nd choice to allow another student to get their 1st,
if that results in a better overall outcome for the group.
"""

import re
import sys
import numpy as np
from scipy.optimize import linear_sum_assignment
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import Counter

# ── CONFIG ───────────────────────────────────────────────────────────────────
INPUT_FILE  = "thesis_data.xlsx"   # default input filename
OUTPUT_FILE = "thesis_allocation_results.xlsx"
HIGH_COST   = 1000  # cost assigned to projects not in a student's preferences
# ─────────────────────────────────────────────────────────────────────────────


def extract_code(text):
    """Extract project code (e.g. 'C037') from a cell value."""
    if not text:
        return None
    m = re.match(r'^(C\d+)', str(text).strip())
    return m.group(1) if m else None


def load_data(filepath):
    print(f"📂 Loading: {filepath}")
    wb = load_workbook(filepath, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))

    # Build project dictionary from all choices in the survey
    projects = {}
    for row in rows[1:]:
        for cell in row[1:]:
            if cell:
                m = re.match(r'^(C\d+)\s+(.*)', str(cell).strip())
                if m and m.group(1) not in projects:
                    projects[m.group(1)] = m.group(2).strip()

    # Parse students and their ranked choices
    students = []
    student_choices = []
    for row in rows[1:]:
        if not row[0]:
            continue
        raw   = str(row[0]).strip()
        parts = raw.split(' ', 1)
        sid   = parts[0]
        sname = parts[1] if len(parts) > 1 else ""
        choices = [extract_code(row[i]) for i in range(1, 9) if i < len(row)]
        choices = [c for c in choices if c]
        students.append((sid, sname))
        student_choices.append(choices)

    print(f"   {len(students)} students | {len(projects)} projects found")
    return students, student_choices, projects


def build_cost_matrix(students, student_choices, projects):
    proj_list = sorted(projects.keys())
    proj_idx  = {p: i for i, p in enumerate(proj_list)}
    n_students = len(students)
    n_projects = len(proj_list)

    cost = np.full((n_students, n_projects), HIGH_COST, dtype=float)
    for s_idx, choices in enumerate(student_choices):
        for rank, proj_code in enumerate(choices, start=1):
            if proj_code in proj_idx:
                cost[s_idx, proj_idx[proj_code]] = rank

    return cost, proj_list


def run_optimisation(cost, proj_list, students, student_choices, projects):
    print("⚙  Running optimisation...")
    row_ind, col_ind = linear_sum_assignment(cost)

    results = []
    assigned_projects = set()

    for s_idx, p_idx in zip(row_ind, col_ind):
        sid, sname  = students[s_idx]
        proj_code   = proj_list[p_idx]
        proj_title  = projects[proj_code]
        choices     = student_choices[s_idx]
        assigned_projects.add(proj_code)

        if proj_code in choices:
            choice_rank = choices.index(proj_code) + 1
            note = ""
        else:
            choice_rank = 99
            note = "⚠ Auto-assigned — not in stated preferences, please review"

        results.append({
            "Student ID":    sid,
            "Student Name":  sname,
            "Project Code":  proj_code,
            "Project Title": proj_title,
            "Choice Rank":   choice_rank,
            "Note":          note
        })

    # Restore original student row order
    order = {sid: i for i, (sid, _) in enumerate(students)}
    results.sort(key=lambda x: order.get(x["Student ID"], 999))

    unassigned = [{"Project Code": c, "Project Title": projects[c]}
                  for c in sorted(projects) if c not in assigned_projects]

    return results, unassigned


def write_results(results, unassigned, output_file):
    rank_dist = Counter(r["Choice Rank"] for r in results)
    auto_count = rank_dist.get(99, 0)

    wb  = Workbook()
    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_ws(ws, hdr_color, alt="EBF3FB"):
        ws.row_dimensions[1].height = 28
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", start_color=hdr_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = bdr
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border    = bdr
                cell.alignment = Alignment(vertical="center")
                if i % 2 == 0:
                    cell.fill = PatternFill("solid", start_color=alt)
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 60)
        ws.freeze_panes = "A2"

    # ── Sheet 1: Allocations ──
    ws1 = wb.active
    ws1.title = "Allocations"
    ws1.append(["Student ID", "Student Name", "Project Code",
                "Project Title", "Choice Rank", "Note"])
    for r in results:
        display_rank = r["Choice Rank"] if r["Choice Rank"] != 99 else "Auto"
        ws1.append([r["Student ID"], r["Student Name"], r["Project Code"],
                    r["Project Title"], display_rank, r["Note"]])
    style_ws(ws1, "1F4E79", "EBF3FB")

    # Colour-code Choice Rank column (col 5)
    for row in ws1.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            v = cell.value
            if v == 1:
                cell.fill = PatternFill("solid", start_color="C6EFCE")
                cell.font = Font(bold=True, color="375623")
            elif v == 2:
                cell.fill = PatternFill("solid", start_color="FFEB9C")
                cell.font = Font(bold=True, color="9C6500")
            elif v == 3:
                cell.fill = PatternFill("solid", start_color="FCE4D6")
                cell.font = Font(bold=True, color="833C00")
            elif isinstance(v, int) and v >= 4:
                cell.fill = PatternFill("solid", start_color="FFC7CE")
                cell.font = Font(bold=True, color="9C0006")
            elif v == "Auto":
                cell.fill = PatternFill("solid", start_color="F4B942")
                cell.font = Font(bold=True, color="7B3F00")

    # ── Sheet 2: Unassigned Projects ──
    ws2 = wb.create_sheet("Unassigned Projects")
    ws2.append(["Project Code", "Project Title"])
    for r in unassigned:
        ws2.append([r["Project Code"], r["Project Title"]])
    style_ws(ws2, "7030A0", "F2EBF9")

    # ── Sheet 3: Summary ──
    ws3 = wb.create_sheet("Summary")
    summary_rows = [
        ("Metric", "Count"),
        ("Total Students Allocated", len(results)),
        ("🟢 Got 1st Choice",  rank_dist.get(1, 0)),
        ("🟡 Got 2nd Choice",  rank_dist.get(2, 0)),
        ("🟠 Got 3rd Choice",  rank_dist.get(3, 0)),
        ("🔴 Got 4th Choice",  rank_dist.get(4, 0)),
        ("🔴 Got 5th Choice",  rank_dist.get(5, 0)),
        ("🔴 Got 6th Choice",  rank_dist.get(6, 0)),
        ("🔴 Got 7th Choice",  rank_dist.get(7, 0)),
        ("🔴 Got 8th Choice",  rank_dist.get(8, 0)),
        ("⚠ Auto-assigned (review needed)", auto_count),
        ("", ""),
        ("Total Projects Available", len(results) + len(unassigned)),
        ("Projects Assigned", len(results)),
        ("Projects Unassigned", len(unassigned)),
        ("", ""),
        ("Run At", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for r in summary_rows:
        ws3.append(r)
    style_ws(ws3, "375623", "F0FFF4")
    for row in ws3.iter_rows(min_row=2, max_col=1):
        if row[0].value:
            row[0].font = Font(bold=True)

    wb.save(output_file)

    # ── Print summary ──
    print(f"\n✅ Done! Results saved to: {output_file}")
    print(f"   Total allocated : {len(results)}")
    print(f"   Got 1st choice  : {rank_dist.get(1,0)}")
    print(f"   Got 2nd choice  : {rank_dist.get(2,0)}")
    print(f"   Got 3rd choice  : {rank_dist.get(3,0)}")
    print(f"   Got 4th+ choice : {sum(v for k,v in rank_dist.items() if isinstance(k,int) and k>=4 and k!=99)}")
    if auto_count:
        print(f"   ⚠  Auto-assigned (review): {auto_count}")


# ── MAIN ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    input_file  = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE
    output_file = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_FILE

    students, student_choices, projects = load_data(input_file)
    cost, proj_list = build_cost_matrix(students, student_choices, projects)
    results, unassigned = run_optimisation(cost, proj_list, students, student_choices, projects)
    write_results(results, unassigned, output_file)
